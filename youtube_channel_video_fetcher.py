import requests
import re
import pandas as pd
import os
import isodate
from openpyxl import load_workbook
from datetime import datetime

# 環境変数からAPIキーを取得
API_KEY = os.getenv("YOUTUBE_API_KEY")
if not API_KEY:
    raise ValueError("環境変数 'YOUTUBE_API_KEY' が設定されていません。")

def get_channel_id(channel_url):
    """チャンネルのURLからチャンネルIDを取得"""
    handle_match = re.search(r"youtube\.com/@([\w-]+)", channel_url)
    if handle_match:
        handle = handle_match.group(1)
        url = f"https://www.googleapis.com/youtube/v3/channels?part=id&forHandle=@{handle}&key={API_KEY}"
    else:
        username_match = re.search(r"youtube\.com/(?:c/|user/)([\w-]+)", channel_url)
        if username_match:
            username = username_match.group(1)
            url = f"https://www.googleapis.com/youtube/v3/channels?part=id&forUsername={username}&key={API_KEY}"
        else:
            channel_id_match = re.search(r"youtube\.com/channel/([\w-]+)", channel_url)
            if channel_id_match:
                return channel_id_match.group(1)
            else:
                print("無効なURLです。")
                return None

    response = requests.get(url).json()
    if "items" in response and response["items"]:
        return response["items"][0]["id"]
    else:
        print("チャンネルIDを取得できませんでした。")
        return None

def format_duration(duration):
    """ISO 8601形式の再生時間を hh:mm:ss 形式に変換"""
    try:
        duration_sec = int(isodate.parse_duration(duration).total_seconds())
        hours, remainder = divmod(duration_sec, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02}:{minutes:02}:{seconds:02}"
    except:
        return "00:00:00"

def get_video_details(video_id):
    """動画の詳細情報を取得"""
    url = f"https://www.googleapis.com/youtube/v3/videos?part=snippet,contentDetails,statistics&id={video_id}&key={API_KEY}"
    response = requests.get(url).json()
    if "items" in response and response["items"]:
        item = response["items"][0]
        title = item["snippet"]["title"]
        published_date = item["snippet"]["publishedAt"][:10]
        duration = format_duration(item["contentDetails"]["duration"])
        view_count = item["statistics"].get("viewCount", "0")
        return title, published_date, duration, int(view_count)
    return None, None, None, None

def get_video_data(channel_id):
    """チャンネルの動画データを取得"""
    url = f"https://www.googleapis.com/youtube/v3/search?key={API_KEY}&channelId={channel_id}&part=id&order=date&maxResults=50"
    videos = []
    while url:
        response = requests.get(url).json()
        for item in response.get("items", []):
            if "videoId" in item["id"]:
                video_id = item["id"]["videoId"]
                title, published_date, duration, view_count = get_video_details(video_id)
                if title:
                    videos.append([title, published_date, duration, view_count, f"https://www.youtube.com/watch?v={video_id}"])
        next_page_token = response.get("nextPageToken")
        if next_page_token:
            url = f"https://www.googleapis.com/youtube/v3/search?key={API_KEY}&channelId={channel_id}&part=id&order=date&maxResults=50&pageToken={next_page_token}"
        else:
            url = None
    return videos

def read_channel_list(filename):
    """検索リストファイルからURLとシート名のペアを取得"""
    wb = load_workbook(filename)
    ws = wb.active
    urls = []
    names = []
    row = 2
    while True:
        url_cell = ws[f"B{row}"].value
        name_cell = ws[f"A{row}"].value
        if not url_cell:
            break
        urls.append(url_cell)
        names.append(str(name_cell))
        row += 1
    return list(zip(names, urls))

def main():
    list_file = "YouTubeチャンネルの動画_検索リスト.xlsx"
    today_str = datetime.now().strftime("%Y%m%d")
    output_filename = f"YouTubeチャンネルの動画データ取得_{today_str}.xlsx"
    
    sheet_data = read_channel_list(list_file)
    writer = pd.ExcelWriter(output_filename, engine="openpyxl")

    for name, url in sheet_data:
        print(f"[処理中] {name}: {url}")
        channel_id = get_channel_id(url)
        if not channel_id:
            print(f"チャンネルIDの取得失敗: {url}")
            continue
        video_data = get_video_data(channel_id)
        if not video_data:
            print(f"動画データなし: {name}")
            continue
        df = pd.DataFrame(video_data, columns=["タイトル", "投稿日", "再生時間", "再生回数", "URL"])
        df.to_excel(writer, sheet_name=name[:31], index=False)  # Excelのシート名制限31文字

    writer.close()
    print(f"[完了] 取得した動画データを {output_filename} に保存しました。")

if __name__ == "__main__":
    main()
